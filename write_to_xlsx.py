"""
version 
@author varlamov.a
@email varlamov.a@rt.ru
@date 06.06.2023
@time 17:13
"""
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font

con = psycopg2.connect(host='localhost', port=5432, user='dq', password='dq', database='dq')

wb = Workbook()
ws1 = wb.active
ws1.title = "Data_from_postgres"

with con.cursor() as cur:
    cur.execute("select * from public.dq_checktask_log limit 10")
    res = cur.description
    cont = cur.fetchall()

col_number = len(cont[0])
line_number = len(cont)
caption_font = Font(bold=True)

for col in range(col_number):
    current_cell = ws1.cell(row=1, column=col + 1, value=res[col].name)
    current_cell.font = caption_font


for line in range(line_number):
    for col in range(col_number):
        ws1.cell(row=line + 2, column=col + 1, value=cont[line][col])

wb.save("example.xlsx")
